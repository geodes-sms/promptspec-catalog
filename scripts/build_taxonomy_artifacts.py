#!/usr/bin/env python3
"""Build taxonomy reduction tables, component matrix, and Excel workbook."""

from __future__ import annotations

import csv
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parents[1]
METHOD = ROOT / "method"
RAW_PATH = METHOD / "master_raw_dataset.csv"
FINAL_PATH = METHOD / "prompt_taxonomy_final_29_patterns.csv"
VARIANTS_PATH = METHOD / "prompt_taxonomy_variants.csv"
MATRIX_PATH = METHOD / "prompt_component_matrix.csv"
WORKBOOK_PATH = (
    METHOD / "prompt_taxonomy_workbook.xlsx"
    if (METHOD / "prompt_taxonomy_workbook.xlsx").exists()
    else METHOD / "taxonomy_reduction.xlsx"
)
COMPONENTS = [
    "Profile/Role",
    "Directive",
    "Context",
    "Procedural Steps",
    "Examples",
    "Output Format/Style",
    "Constraints",
]
MERGED_STATUS = "MERGED_CONSOLIDATION"


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_rows(path: Path, fields: list[str], rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def canonical_status(rows: list[dict[str, str]]) -> str:
    if any(row["status"] == "INCLUDED" for row in rows):
        return "INCLUDED"
    excluded = {
        row["status"]
        for row in rows
        if row["status"] != MERGED_STATUS
    }
    if len(excluded) != 1:
        raise ValueError(f"canonical status is ambiguous: {sorted(excluded)}")
    return next(iter(excluded))


def concept_rows(raw_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in raw_rows:
        grouped[row["canonical_key"]].append(row)
    concepts = []
    for key in sorted(grouped):
        rows = grouped[key]
        status = canonical_status(rows)
        canonical_rows = [
            row for row in rows if row["status"] != MERGED_STATUS
        ]
        representative = next(
            (row for row in canonical_rows if row["status"] == status),
            canonical_rows[0],
        )
        concepts.append(
            {
                "canonical_key": key,
                "raw_names": " | ".join(sorted({row["raw_name"] for row in rows})),
                "sources": " | ".join(sorted({row["source_key"] for row in rows})),
                "status": status,
                "decision_class": representative["decision_class"],
                "folds_into": representative["folds_into"],
                "reason": representative["status_reason"],
            }
        )
    return concepts


def build_variants(raw_rows: list[dict[str, str]]) -> None:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in raw_rows:
        if row.get("decision_class") == "variant":
            grouped[(row.get("merged_from_key") or "").strip()].append(row)

    rows = []
    for merged_from_key in sorted(grouped):
        source_rows = grouped[merged_from_key]
        rows.append(
            {
                "canonical_key": merged_from_key,
                "parent_canonical_key": source_rows[0]["canonical_key"],
                "raw_names": " | ".join(
                    sorted({row["raw_name"] for row in source_rows})
                ),
                "stage": "consolidation",
                "folds_into": source_rows[0]["folds_into"],
                "reason": source_rows[0]["status_reason"],
            }
        )
    write_rows(
        VARIANTS_PATH,
        [
            "canonical_key",
            "parent_canonical_key",
            "raw_names",
            "stage",
            "folds_into",
            "reason",
        ],
        rows,
    )


def build_matrix(final_rows: list[dict[str, str]]) -> None:
    rows = []
    for row in final_rows:
        selected = {
            value.strip() for value in row["Component use"].split(",") if value.strip()
        }
        matrix_row: dict[str, object] = {
            "Pattern Name": row["Pattern Name"],
            "Category": row["Category"],
            "Subcategory": row["Subcategory"],
        }
        matrix_row.update({component: int(component in selected) for component in COMPONENTS})
        rows.append(matrix_row)
    write_rows(
        MATRIX_PATH,
        ["Pattern Name", "Category", "Subcategory", *COMPONENTS],
        rows,
    )


def append_table(sheet, headers: list[str], rows: list[list[object]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        sheet.column_dimensions[column[0].column_letter].width = width


def build_workbook(
    raw_rows: list[dict[str, str]],
    final_rows: list[dict[str, str]],
    concepts: list[dict[str, str]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    source_counts = Counter(row["source_key"] for row in raw_rows)
    normalized_by_source = {
        source: len(
            {row["normalized_name"] for row in raw_rows if row["source_key"] == source}
        )
        for source in source_counts
    }
    canonical_by_source = {
        source: len(
            {
                row["canonical_key"]
                for row in raw_rows
                if row["source_key"] == source
            }
        )
        for source in source_counts
    }
    spine = workbook.create_sheet("Spine")
    append_table(
        spine,
        ["Scope", "Raw rows", "Unique normalized names", "Distinct concepts"],
        [
            [
                source,
                source_counts[source],
                normalized_by_source[source],
                canonical_by_source[source],
            ]
            for source in sorted(source_counts)
        ]
        + [
            [
                "TOTAL",
                len(raw_rows),
                len({row["normalized_name"] for row in raw_rows}),
                len({row["canonical_key"] for row in raw_rows}),
            ]
        ],
    )

    retained = workbook.create_sheet("Retained")
    append_table(
        retained,
        ["Pattern Name", "Category", "Subcategory", "Component use"],
        [
            [
                row["Pattern Name"],
                row["Category"],
                row["Subcategory"],
                row["Component use"],
            ]
            for row in final_rows
        ],
    )

    non_retained = workbook.create_sheet("Non-retained")
    excluded = [
        row for row in concepts if row["status"] == "EXCLUDED_OUT_OF_SCOPE"
    ]
    append_table(
        non_retained,
        [
            "canonical_key",
            "raw_names",
            "sources",
            "status",
            "decision_class",
            "folds_into",
            "reason",
        ],
        [
            [
                row["canonical_key"],
                row["raw_names"],
                row["sources"],
                row["status"],
                row["decision_class"],
                row["folds_into"],
                row["reason"],
            ]
            for row in excluded
        ],
    )

    counts = workbook.create_sheet("Counts")
    status_counts = Counter(row["status"] for row in concepts)
    decision_counts = Counter(
        row["decision_class"]
        for row in concepts
        if row["status"] == "EXCLUDED_OUT_OF_SCOPE"
    )
    normalized_count = len({row["normalized_name"] for row in raw_rows})
    merged_count = len(
        {
            (row.get("merged_from_key") or "").strip()
            for row in raw_rows
            if row.get("decision_class") == "variant"
        }
    )
    post_consolidation_count = len(concepts)
    consolidation_merges = normalized_count - post_consolidation_count
    count_rows = [
        ["spine", "raw rows", len(raw_rows)],
        ["spine", "unique normalized names", normalized_count],
        ["spine", "consolidation merges", consolidation_merges],
        ["spine", "distinct concepts after consolidation", post_consolidation_count],
        ["status", "INCLUDED", status_counts["INCLUDED"]],
        ["status", "EXCLUDED_WORKFLOW", status_counts["EXCLUDED_WORKFLOW"]],
        [
            "status",
            "EXCLUDED_OUT_OF_SCOPE (non-retained)",
            status_counts["EXCLUDED_OUT_OF_SCOPE"],
        ],
        ["consolidation", "variant folds", merged_count],
    ]
    count_rows.extend(
        ["non-retained decision_class", key, decision_counts[key]]
        for key in sorted(decision_counts)
    )
    append_table(counts, ["Bucket type", "Bucket", "Distinct concepts"], count_rows)

    workbook.save(WORKBOOK_PATH)


def main() -> None:
    raw_rows = read_rows(RAW_PATH)
    final_rows = read_rows(FINAL_PATH)
    concepts = concept_rows(raw_rows)
    build_variants(raw_rows)
    build_matrix(final_rows)
    build_workbook(raw_rows, final_rows, concepts)
    print(
        "Built "
        f"{VARIANTS_PATH.relative_to(ROOT)}, {MATRIX_PATH.relative_to(ROOT)}, "
        f"and {WORKBOOK_PATH.relative_to(ROOT)}"
    )


if __name__ == "__main__":
    main()
